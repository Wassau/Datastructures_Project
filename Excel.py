from openpyxl import load_workbook

FILE_PATH = 'Test.xlsx'
SHEET = 'Hoja1'
workbook = load_workbook(FILE_PATH, read_only = False)
sheet = workbook[SHEET]

for id_user, nombre, apellido, saturacion, frecuencia_car, hora_lect, fecha_lect in sheet.iter_rows():
    print(str(id_user.value) + " " + nombre.value + " " + apellido.value + " " + str(saturacion.value) + " " + str(frecuencia_car.value) + " " + str(hora_lect.value) + " " + str(fecha_lect.value))
    
